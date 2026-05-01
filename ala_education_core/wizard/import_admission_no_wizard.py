import base64
import io
import openpyxl

from odoo import models, fields, _
from odoo.exceptions import UserError


class ImportAdmissionNoWizard(models.TransientModel):
    _name = 'import.admission.no.wizard'
    _description = 'Import Admission Number'

    file = fields.Binary(string="Excel File", required=True)
    file_name = fields.Char(string="File Name")

    def action_import_admission_no(self):
        if not self.file:
            raise UserError(_("Please upload Excel file."))

        # 🔥 Normalize function (remove spaces + lowercase)
        def normalize(value):
            return ''.join(str(value or '').split()).lower()

        try:
            data = base64.b64decode(self.file)
            workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Invalid Excel file: %s") % e)

        # Read headers
        headers = {}
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col).value
            if value:
                headers[str(value).strip().lower()] = col

        required_cols = ['name', 'registration number', 'admission no']
        for col in required_cols:
            if col not in headers:
                raise UserError(_("Missing column in Excel: %s") % col)

        Student = self.env['ala.education.student'].sudo()

        updated = 0
        skipped = 0
        not_found = []

        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row=row, column=headers['name']).value
            registration_no = sheet.cell(row=row, column=headers['registration number']).value
            admission_no = sheet.cell(row=row, column=headers['admission no']).value

            # Skip empty rows
            if not name or not registration_no or not admission_no:
                skipped += 1
                continue

            name = str(name).strip()
            registration_no = str(registration_no).strip()
            admission_no = self.format_admission_no(admission_no)

            name_clean = normalize(name)
            reg_clean = normalize(registration_no)

            # 🔍 Step 1: search loosely
            students = Student.search([
                ('name', 'ilike', name[:5])  # partial search for performance
            ])

            if not students:
                not_found.append(f"Row {row}: Name not found - {name}")
                continue

            # 🔍 Step 2: strict match (ignore spaces + case)
            student = students.filtered(
                lambda s:
                normalize(s.name) == name_clean and
                normalize(s.register_no) == reg_clean
            )

            if not student:
                not_found.append(
                    f"Row {row}: Not matched - {name} / {registration_no}"
                )
                continue

            student_rec = student[0]

            before_value = student_rec.admission_no

            # ✅ Update field (IMPORTANT: change to ad_no if needed)
            student_rec.write({
                'admission_no': admission_no
            })

            student_rec.invalidate_recordset(['admission_no'])

            after_value = student_rec.admission_no

            print(
                "UPDATED:",
                f"Row={row}",
                f"ID={student_rec.id}",
                f"Name={student_rec.name}",
                f"RegNo={student_rec.register_no}",
                f"Before={before_value}",
                f"Excel={admission_no}",
                f"After={after_value}"
            )

            updated += 1

        # Final message
        message = (
            f"✅ Admission No Import Completed\n\n"
            f"✔ Updated: {updated}\n"
            f"⏭ Skipped: {skipped}\n"
            f"❌ Not Matched: {len(not_found)}"
        )

        if not_found:
            message += "\n\nDetails:\n" + "\n".join(not_found[:50])

        raise UserError(_(message))

    def format_admission_no(value):
        value = str(value or '').strip()

        if '/' in value:
            left, right = value.rsplit('/', 1)
            right = right.strip()

            if len(right) == 2 and right.isdigit():
                value = "%s/20%s" % (left.strip(), right)

        return value